import time

import sys
import os
import shutil
import traceback

import utility as u
from pb_shell import pb_shell
import openpyxl
from templite import Templite

from unreal_import_switch import unreal


def exclude_proto_names():
    return [
        'options_ext',
        'excel_basic',
    ]

def get_pb_field_ref(excel_field_name):
    """
    根据excel列描述名称[第二列的字符串]，生成对应的pb message字段映射，这里处理了 '.' 操作获得子Message以及 '[]' 操作获得repeated字段按下标索引

    @excel_field_name (str)
      excel列描述名称
    """
    pb_field_excel_ref = []

    for section in excel_field_name.split("."):
        start_of_array_index = section.find("[")
        name = section
        array_index = None

        if start_of_array_index > 0:
            array_index = int(section[start_of_array_index + 1:-1])
            name = section[0:start_of_array_index]

        pb_field_excel_ref.append((name, array_index))
    return pb_field_excel_ref


def validate(excel_instances):
    # ** TODO:重复主键检查等
    return excel_instances != None


def output_binaries(excel_instances, output_path, type_name):
    """
    将序列化的内容输出到指定的文件路径

    @excel_instances (list(object<Message>))
      已序列化对象的列表
    @output_path (str)
      输出文件的完整路径
    @type_name (str)
      表名
    """
    if not os.path.exists(output_path):
        os.makedirs(output_path)
    # ** 构造excel的包装类，将实例数据都填充到rows字段中
    from pb_helper import pb_helper
    excel_wrap_type, _ = pb_helper.get_type_with_module('Excel')
    if excel_wrap_type is None:
        unreal.log_error('wrap type "{0}" not found. failed to output excel'.format(excel_wrap_type))
        sys.exit(-1)
    excel_wrap = excel_wrap_type()
    setattr(excel_wrap, 'excelName', type_name)
    rows = getattr(excel_wrap, 'rows')
    for excel_instance in excel_instances:
        rows_data = excel_instance.SerializeToString()
        rows.append(rows_data)

    # ** 将包装类整个写入到文件
    write_file = os.path.join(output_path, "{}".format(type_name))
    with open(write_file, "wb") as f:
        f.write(excel_wrap.SerializeToString())


def excel_binarization(excel, type_name, verbose=True):
    """
    序列化指定的excel中每一行到指定的pb类型

    @excel (str)
      excel文件的完整路径
    @type_name (str)
      对应的pb message名称
    @verbose (bool)
      是否显示序列化log信息
    """
    from pb_helper import pb_helper
    pb_type, _ = pb_helper.get_type_with_module(type_name)
    result = []
    if pb_type is None:
        unreal.log_error('pb type "{0}" not found. failed to binarization excel "{1}"'.format(type_name, excel))
        return

    workbook = openpyxl.load_workbook(excel)
    worksheet = workbook.worksheets[0]
    current_row = 1
    current_columns = 1

    try:
        if verbose:
            unreal.log("excel binarization => {}".format(excel))
            start_time = time.time()
        # ** NOTE:列所指代的特定pb字段，注意这里是可以使用伸缩格式（即fieldA.fieldB.fieldC）
        columns_to_pb_field_excel_ref = {}
        # ** 先记录数据档第二行对应的pb字段
        for columns in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=2, column=columns).value
            if type(cell_value) != str:
                continue
            excel_field_name = cell_value.strip()
            if len(excel_field_name) > 0:
                pb_field_excel_ref = get_pb_field_ref(excel_field_name)
                columns_to_pb_field_excel_ref[columns] = pb_field_excel_ref

        # ** 从第三行开始
        for rows in range(3, worksheet.max_row + 1):
            current_row = rows
            # ** 创建指定pb类型的实例，一行对应一个实例
            shell = pb_shell(pb_type)
            for columns in range(1, worksheet.max_column + 1):
                current_columns = columns
                if columns in columns_to_pb_field_excel_ref:
                    sheet_cell = worksheet.cell(row=rows, column=columns)
                    cell_value = sheet_cell.value
                    # ** NOTE:暂时跳过公式的处理，后面如果需要代入公式再看怎么做
                    if sheet_cell.data_type == "f":
                        continue
                    if type(cell_value) == str:
                        cell_value = cell_value.strip()
                    shell.write_field_value(columns_to_pb_field_excel_ref[columns], cell_value)
            if shell.valid and shell.instance.IsInitialized():
                result.append(shell.instance)

    except Exception as e:
        a_num = int.from_bytes(b"A", byteorder="little") - 1
        second_alphabet = current_columns // 26
        first_alphabet = (current_columns - 1) % 26 + 1
        if second_alphabet > 0:
            out_bytes = bytes([second_alphabet + a_num, first_alphabet + a_num])
        else:
            out_bytes = bytes([first_alphabet + a_num])

        unreal.log_error("{} binarization failed in [line {} Columns {}]".format(excel, current_row, str(out_bytes, encoding="utf-8")))
        unreal.log_error(traceback.format_exc())
        raise e

    if verbose:
        unreal.log("<= {0} binarization finished in {1:.2f} seconds".format(excel, time.time() - start_time))

    return result


def get_bin():
    """
    根据平台，获得protoc执行文件路径
    """
    if u.is_win():
        project_plugins_dir = unreal.Paths.project_plugins_dir()
        result = os.path.abspath(os.path.normpath(os.path.join(project_plugins_dir, "Protobuf/ThirdParty/protobuf/bin/protoc.exe")))
        return f"\"{result}\""


def make_args(dir, ignore_patterns=None):
    """
    生成 protoc 执行时需要的 .proto 文件对应的路径参数

    @dir (str)
      proto文件所在根目录
    @ignore_patterns (list(str))
      参数解释
    """
    proto_files = u.get_files(dir, ["*.proto"], ignore_patterns)

    # ** NOTE:省略了比较时间戳和过滤生成
    if len(proto_files) > 0:
        project_plugins_dir = unreal.Paths.project_plugins_dir()
        include_path = os.path.abspath(os.path.normpath(os.path.join(project_plugins_dir, "Protobuf/Content/Proto/")))
        args = [
            f"--proto_path=\"{dir}\"",
            f"--proto_path=\"{include_path}\"",
        ]
        proto_files.extend(u.get_files(include_path, ["*.proto", ignore_patterns]))
        for proto_file in proto_files:
            args.append(f"\"{proto_file}\"")
        return args
    else:
        unreal.log_warning("No proto files in {} ??".format(dir))
    return None


def generate_pbdef(proto_path, output_path, type):
    if os.path.exists(output_path):
        shutil.rmtree(output_path)
    os.makedirs(output_path)
    if type == "python":
        args = make_args(proto_path)
        args.append(f"--python_out=\"{output_path}\"")
    elif type == "cpp":
        args = make_args(proto_path, ["*[\\/]google[\\/]*"])
        args.append(f"--cpp_out=\"{output_path}\"")
    else:
        unreal.log_warning("not supported type {}".format(type))
        return
    u.execute(get_bin(), args)


def filter_by_option_descriptor(pb_fields, option_field_descriptor_name):
    """
    从输出的pb fields中，筛选出含有指定pb option的pb field

    @pb_fields (list(object<FieldDescriptor>))
      待筛选的 Message FieldDescriptor 对象列表
    @option_field_descriptor_name (str)
      指定的 option
    """

    result_fields = []

    for pb_field in pb_fields:
        is_target_field = False
        if pb_field.has_options:
            options = pb_field.GetOptions()
            for option_field_descriptor in options.Extensions:
                if option_field_descriptor.name == option_field_descriptor_name:
                    is_target_field = True
        if is_target_field:
            result_fields.append(pb_field)
    return result_fields


def get_option_value(options, target_option, default=''):
    """
    从options结构中获得target_option的值，如果找不到返回空字符串

    @options (object)
      options字段，是DESCRIPTOR中的一个结构，详情见google\protobuf\descriptor.py

    @target_option (str)
      option名称，参考options_ext.proto定义
    
    @default (str)
      找不到时需要返回的默认值
    """
    for option_field_descriptor, option_field_value in options.ListFields():
        if option_field_descriptor.name == target_option:
            return option_field_value
    return default


def handle_friend_class(cpp_wrapper_content_by_modules):
    """
    处理友元类

    @cpp_wrapper_content_by_modules (dict(str, dict()))
      已加载并处理好的全部模块映射数据
    """

    all_class_wrapper = []
    all_excel_wrapper = []
    for loaded_module in cpp_wrapper_content_by_modules:
        cpp_wrapper_content = cpp_wrapper_content_by_modules[loaded_module]
        all_class_wrapper.extend(cpp_wrapper_content['classes_wrapper'])
        all_excel_wrapper.extend(cpp_wrapper_content['excels_wrapper'])

    for class_wrapper in all_class_wrapper:
        class_friend_classes = [] if 'class_friend_classes' not in class_wrapper else class_wrapper['class_friend_classes']
        # 自身Message的Excel包装能够使用自身的Load
        for excel_wrapper in all_excel_wrapper:
            if excel_wrapper["excelname"] == class_wrapper["name"]:
                class_friend_classes.append(f'{excel_wrapper["excelname"]}Excel')
        # 自身能使用子Message中的Load
        for pb_field in class_wrapper['pb_fields']:
            if pb_field.message_type is not None:
                for inner_class_wrapper in all_class_wrapper:
                    if inner_class_wrapper['name'] == pb_field.message_type.name:
                        inner_class_friend_classes = [] if 'class_friend_classes' not in inner_class_wrapper else inner_class_wrapper['class_friend_classes']
                        inner_class_friend_classes.append(f'{class_wrapper["name"]}{class_wrapper["typename_postfix"]}')
                        inner_class_wrapper['class_friend_classes'] = inner_class_friend_classes
        class_wrapper['class_friend_classes'] = class_friend_classes


def generate_cpp_wrapper(output_path, cpp_excel_wrapper):
    """
    生成UE的cpp包装，将excel对应的Message全部转换到UCLASS，这部分cpp代码是靠templite模板生成的

    @output_path (str)
      excel文件的完整路径
    @cpp_excel_wrapper (str)
      对应的pb message名称
    """
    cpp_wrapper_content_by_modules = {}
    from pb_helper import pb_helper
    preference = pb_helper.load_proto_gen_preference()
    uclass_as_default = preference["uclass_as_default"] == 'True'
    struct_typename_postfix = preference["struct_typename_postfix"]
    class_typename_postfix = preference["class_typename_postfix"]
    excel_typename_postfix = preference["excel_typename_postfix"]
    for loaded_module in pb_helper.loaded_modules:
        module = sys.modules[loaded_module]
        # ** NOTE:仅windows生效
        if 'google\\protobuf' in module.__file__:
            continue
        if loaded_module.replace('_pb2', '') in exclude_proto_names():
            continue
        pb_imports = []
        for dependency in module.DESCRIPTOR.dependencies:
            shortname = dependency.name.replace('.proto', '')
            if shortname in exclude_proto_names():
                continue
            pb_imports.append(shortname)

        # ** NOTE:根据模块将类写入对应wrapper中
        classes_wrapper = []
        structs_wrapper = []
        for message_name in module.DESCRIPTOR.message_types_by_name:
            pb_type = getattr(module, message_name)
            pb_fields = pb_type.DESCRIPTOR.fields
            option_FSoftClassPath_fields = filter_by_option_descriptor(pb_fields, 'softclasspath')
            option_FSoftObjectPath_fields = filter_by_option_descriptor(pb_fields, 'softobjectpath')
                
            option_pk_fields = filter_by_option_descriptor(pb_fields, 'pk')

            options = pb_type.DESCRIPTOR.GetOptions()
            ustruct_specifiers = get_option_value(options, 'ustruct', False)
            uclass_specifiers = get_option_value(options, 'uclass')

            if not uclass_as_default and ustruct_specifiers == False:
                ustruct_specifiers = ''
            # ** 生成USTRUCT包装
            struct_wrapper = {
                'name': message_name,
                'pb_fields': pb_fields,
                'option_FSoftClassPath_fields': option_FSoftClassPath_fields,
                'option_FSoftObjectPath_fields': option_FSoftObjectPath_fields,
                'option_pk_fields': option_pk_fields,
                'ustruct_specifiers': ustruct_specifiers if ustruct_specifiers != False else '',
                'typename_postfix' : struct_typename_postfix,
            }
            #BlueprintType default, 为了让Excel的Rows能够BlueprintReadOnly
            if 'BlueprintType' not in struct_wrapper['ustruct_specifiers']:
                if struct_wrapper['ustruct_specifiers'] == '':
                    struct_wrapper['ustruct_specifiers'] = 'BlueprintType'
                else:
                    struct_wrapper['ustruct_specifiers'] += ', BlueprintType'
            # ** 生成UCLASS包装
            class_wrapper = {
                'name': message_name,
                'pb_fields': pb_fields,
                'option_FSoftClassPath_fields': option_FSoftClassPath_fields,
                'option_FSoftObjectPath_fields': option_FSoftObjectPath_fields,
                'option_pk_fields': option_pk_fields,
                'uclass_specifiers': uclass_specifiers,
                'is_ustruct': ustruct_specifiers != False,
                'typename_postfix' : class_typename_postfix,
                'struct_wrapper' : struct_wrapper
            }
            if not uclass_as_default or ustruct_specifiers != False:
                structs_wrapper.append(struct_wrapper)

            classes_wrapper.append(class_wrapper)

        enums_wrapper = []
        for enum_name in module.DESCRIPTOR.enum_types_by_name:
            pb_enum = getattr(module, enum_name)
            options = pb_enum.DESCRIPTOR.GetOptions()
            enum_wrapper = {
                'enumname': enum_name,
                'pb_enum_values': pb_enum.DESCRIPTOR.values,
                'uenum_specifiers': get_option_value(options, 'uenum')
            }
            enums_wrapper.append(enum_wrapper)

        # ** NOTE:模板中会根据以下数据生成对应的cpp类，其中：
        # ** classes_wrapper生成class是对pb中Message定义的包装，
        # ** enums_wrapper生成enum是对pb中Enum定义的包装，
        # ** excels_wrapper生成class是对一个excel表的包装，仅对应有Excel表的Message才会生成这个class，这个是延后添加的，
        # ** pb_imports生成文件中依赖的其他生成文件，依据pb import依赖关系
        cpp_wrapper_content_by_modules[loaded_module] = {
            'classes_wrapper': classes_wrapper,
            'structs_wrapper': structs_wrapper,
            'enums_wrapper': enums_wrapper,
            'excels_wrapper': [],
            'pb_imports': pb_imports,
        }

    for excel_wrapper in cpp_excel_wrapper:
        pb_type, loaded_module = pb_helper.get_type_with_module(excel_wrapper)
        if pb_type is None:
            # ** 跳过非本模块的excel定义
            continue
        cpp_wrapper_content = cpp_wrapper_content_by_modules[loaded_module]
        classes_wrapper = cpp_wrapper_content['classes_wrapper']
        # ** 找到Excel对应的class包装，直接从中获取Excel包装需要的数据
        excel_class_wrapper = None
        for class_wrapper in classes_wrapper:
            if class_wrapper['name'] == excel_wrapper:
                excel_class_wrapper = class_wrapper
                break
        excels_wrapper = cpp_wrapper_content['excels_wrapper']
        excel_wrapper = {
            'excelname': excel_wrapper,
            'pb_fields': excel_class_wrapper['pb_fields'],
            'option_pk_fields': excel_class_wrapper['option_pk_fields'],
            'typename_postfix': excel_typename_postfix,
            'class_wrapper' : excel_class_wrapper,
            'is_ustruct': excel_class_wrapper['is_ustruct'],
        }
        excels_wrapper.append(excel_wrapper)

    from template.cpp_wrapper import cpp_wrapper_template, pb_type_to_ue_type_map
    if not os.path.exists(output_path):
        os.makedirs(output_path)

    handle_friend_class(cpp_wrapper_content_by_modules)

    for loaded_module in cpp_wrapper_content_by_modules:
        cpp_wrapper_content = cpp_wrapper_content_by_modules[loaded_module]
        module_name = loaded_module.replace('_pb2', '')

        # ** NOTE:对每一个Proto文件中的所有Message生成一个对应的UCLASS包装
        file_basename = '{}Wrapper'.format(module_name)
        file_path = os.path.join(output_path, '{}.h'.format(file_basename))
        if os.path.exists(file_path):
            os.remove(file_path)

        cpp_wrapper_content = Templite(cpp_wrapper_template).render(
            classes_wrapper=cpp_wrapper_content['classes_wrapper'],
            structs_wrapper=cpp_wrapper_content['structs_wrapper'],
            enums_wrapper=cpp_wrapper_content['enums_wrapper'],
            excels_wrapper=cpp_wrapper_content['excels_wrapper'],
            pb_imports=cpp_wrapper_content['pb_imports'],
            module_name=module_name,
            file_basename=file_basename,
            pb_type_to_ue_type_map=pb_type_to_ue_type_map,
            FieldDescriptor=pb_helper.FieldDescriptor,
            get_option_value=get_option_value,
            uclass_as_default=uclass_as_default
        )

        with open(file_path, 'w') as f:
            f.write(cpp_wrapper_content)


def generate_excel_bin():
    """
    根据已生成的python pb和对应excel生成二进制pb数据
    返回成功生成二进制pb数据对应的excel名称

    """
    from pb_helper import pb_helper

    preference = pb_helper.load_protobuf_preference()

    excel_path = pb_helper.get_path_from_preference(preference, 'excel_root_path')
    binaries_out = pb_helper.get_path_from_preference(preference, 'binaries_out')

    excel_files = u.get_files(excel_path, ["*.xlsx", "*.xlsm"], ["*~$*.xlsx"], recursive=False)
    cpp_excel_wrapper = []
    for excel_file in excel_files:
        basename, _ = os.path.splitext(os.path.basename(excel_file))
        excel_instances = excel_binarization(excel_file, basename)
        if validate(excel_instances):
            output_binaries(excel_instances, binaries_out, basename)
            cpp_excel_wrapper.append(basename)

    return cpp_excel_wrapper

def generate_all():
    """
    生成所有pb相关内容，包括excel对应二进制pb数据、message对应cpp包装等

    """
    from pb_helper import pb_helper

    preference = pb_helper.load_protobuf_preference()

    proto_path = pb_helper.get_path_from_preference(preference, 'proto_root_path')
    cpp_proto_out = pb_helper.get_path_from_preference(preference, 'cpp_proto_out')
    ue_cpp_wrapper_out = pb_helper.get_path_from_preference(preference, 'ue_cpp_wrapper_out')

    python_pb_out = pb_helper.get_pb_path()
    unreal.log("python_pbdef output path : {}".format(python_pb_out))
    generate_pbdef(proto_path, python_pb_out, 'python')
    generate_pbdef(proto_path, cpp_proto_out, "cpp")

    cpp_excel_wrapper = generate_excel_bin()
    generate_cpp_wrapper(ue_cpp_wrapper_out, cpp_excel_wrapper)
